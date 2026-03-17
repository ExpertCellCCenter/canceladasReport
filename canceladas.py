import streamlit as st
import pandas as pd
import pyodbc
from datetime import date
from io import BytesIO
from openpyxl.utils import get_column_letter

# -------------------------------------------------
# CONFIGURACIÓN
# -------------------------------------------------
st.set_page_config(page_title="Dashboard Canceladas", layout="wide")

def get_connection():
    cfg = st.secrets["db"]
    conn_str = (
        f"DRIVER={{{cfg['driver']}}};"
        f"SERVER={cfg['server']};"
        f"DATABASE={cfg['database']};"
        f"UID={cfg['username']};"
        f"PWD={cfg['password']};"
        "Encrypt=yes;TrustServerCertificate=yes;"
    )
    return pyodbc.connect(conn_str, autocommit=True)

def df_to_excel_bytes(df: pd.DataFrame, sheet_name: str = "Canceladas") -> bytes:
    output = BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name=sheet_name)
        ws = writer.sheets[sheet_name]
        
        max_row = ws.max_row
        max_col = ws.max_column
        if max_col > 0 and max_row > 0:
            ws.auto_filter.ref = f"A1:{get_column_letter(max_col)}{max_row}"
            for col_idx in range(1, max_col + 1):
                col_letter = get_column_letter(col_idx)
                max_length = 0
                for cell in ws[col_letter]:
                    if cell.value is not None:
                        max_length = max(max_length, len(str(cell.value)))
                ws.column_dimensions[col_letter].width = max_length + 2
                
    output.seek(0)
    return output.getvalue()

# -------------------------------------------------
# EXTRACCIÓN Y TRANSFORMACIÓN
# -------------------------------------------------
@st.cache_data
def load_supervisores():
    """Carga el catálogo de empleados para obtener al Jefe Inmediato (Supervisor)"""
    sql = """
    SELECT DISTINCT
        e.[Nombre Completo] AS NombreCompleto,
        e.[Jefe Inmediato]  AS JefeDirecto
    FROM reporte_empleado('EMPRESA_MAESTRA',1,'','') AS e
    WHERE
        e.[Canal de Venta] = 'ATT'
        AND e.[Operacion]   = 'CONTACT CENTER'
        AND e.[Tipo Tienda] = 'VIRTUAL'
        AND e.[Estatus] = 'ACTIVO';
    """
    conn = get_connection()
    df = pd.read_sql(sql, conn)
    conn.close()
    
    # Limpieza básica para el jefe directo
    df["JefeDirecto"] = df["JefeDirecto"].astype(str).str.strip()
    df["JefeDirecto"] = df["JefeDirecto"].replace({"nan": "", "None": ""})
    df["JefeDirecto"] = df["JefeDirecto"].replace("", "ENCUBADORA")
    
    return df

@st.cache_data
def load_canceladas(fecha_ini, fecha_fin):
    fi = fecha_ini.strftime("%Y%m%d")
    ff = fecha_fin.strftime("%Y%m%d")
    
    sql = f"""
    SELECT *
    FROM reporte_programacion_entrega('empresa_maestra', 4, '{fi}', '{ff}')
    WHERE [Estatus] = 'Canc Error';
    """
    
    conn = get_connection()
    cur = conn.cursor()
    cur.execute("SET NOCOUNT ON; SET ANSI_WARNINGS OFF;")
    df = pd.read_sql(sql, conn)
    cur.execute("SET ANSI_WARNINGS ON;")
    cur.close()
    conn.close()
    
    return df

def calcular_estatus_anterior(row):
    """Calcula el último estatus antes de la cancelación basándose en las fechas"""
    fechas = {}
    
    if pd.notna(row.get('Fecha creacion')): 
        fechas['Nuevo'] = pd.to_datetime(row['Fecha creacion'], errors='coerce')
    if pd.notna(row.get('Back Office')): 
        fechas['Back Office'] = pd.to_datetime(row['Back Office'], errors='coerce')
    if pd.notna(row.get('Solicitado')): 
        fechas['Solicitado'] = pd.to_datetime(row['Solicitado'], errors='coerce')
    if pd.notna(row.get('Entregado')): 
        fechas['Entregado'] = pd.to_datetime(row['Entregado'], errors='coerce')
    
    fechas_validas = {k: v for k, v in fechas.items() if pd.notna(v) and v.year > 1900}
    
    if not fechas_validas:
        return "Desconocido"
        
    estatus_anterior = max(fechas_validas, key=fechas_validas.get)
    return estatus_anterior

# -------------------------------------------------
# INTERFAZ PRINCIPAL
# -------------------------------------------------
st.title("🚫 Dashboard: Órdenes Canceladas")

with st.sidebar:
    st.header("Filtros Temporales")
    
    today = date.today()
    first_of_march = date(today.year, 3, 1)
    
    f_ini = st.date_input("Fecha Inicio", first_of_march)
    f_fin = st.date_input("Fecha Fin", today)
    
    if st.button("🔄 Actualizar datos"):
        st.cache_data.clear()
        st.rerun()

# 1. Cargar ambas bases de datos
df_raw = load_canceladas(f_ini, f_fin)
df_empleados = load_supervisores()

if df_raw.empty:
    st.info("No hay cancelaciones ('Canc Error') en este periodo.")
else:
    df_procesado = df_raw.copy()
    
    # 2. Hacer el cruce (JOIN) para traer al Supervisor
    df_procesado = df_procesado.merge(
        df_empleados,
        how="left",
        left_on="Vendedor",
        right_on="NombreCompleto"
    )
    
    # Llenar vacíos por si algún vendedor no tiene supervisor asignado
    df_procesado["JefeDirecto"] = df_procesado["JefeDirecto"].fillna("SIN SUPERVISOR")
    
    # 3. Asignar usuario de cancelación
    if 'Usuario cancleacion' in df_procesado.columns:
        df_procesado['Log_Cancelacion (Usuario)'] = df_procesado['Usuario cancleacion']
    else:
        df_procesado['Log_Cancelacion (Usuario)'] = "No disponible"
        
    # 4. Calcular el estatus anterior
    df_procesado['Log_Anterior (Estatus)'] = df_procesado.apply(calcular_estatus_anterior, axis=1)
    
    # Renombrar columnas para la tabla final
    df_procesado = df_procesado.rename(columns={
        "JefeDirecto": "Supervisor",
        "Vendedor": "Ejecutivo", 
        "Fecha cancelacion": "Fecha Cancelación"
    })
            
    # 5. Seleccionar y ordenar las columnas a mostrar
    cols_display = [
        "Folio", "Supervisor", "Ejecutivo", "Cliente", "Estatus", 
        "Fecha Cancelación", "Log_Anterior (Estatus)", "Log_Cancelacion (Usuario)"
    ]
    
    # Validar que existan antes de mostrarlas
    cols_display = [c for c in cols_display if c in df_procesado.columns]
    df_display = df_procesado[cols_display]

    # Métricas y Tabla
    st.metric("Total Canceladas", len(df_display))
    st.dataframe(df_display, use_container_width=True, hide_index=True)
    
    # Descarga
    st.download_button(
        label="📥 Descargar Canceladas en Excel",
        data=df_to_excel_bytes(df_display, "Canceladas"),
        file_name=f"Canceladas_{f_ini}_al_{f_fin}.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
